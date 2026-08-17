from flask import Flask, render_template, request, redirect, url_for, Response, flash, jsonify, session, get_flashed_messages
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, and_, desc, func
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from functools import wraps
from datetime import datetime, timedelta
import csv
import io
import os
import re

load_dotenv()

app = Flask(__name__)

secret_key = os.getenv('SECRET_KEY', 'soiltrack_secret_key_2026_super_secure_987')
app.secret_key = secret_key

# Task 4 & Task 5: CSRF protection and 30-minute session timeout
csrf = CSRFProtect(app)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Database Configuration (Supports Cloud MySQL, Railway, Render, PythonAnywhere & SQLite fallback)
db_url = os.getenv('DATABASE_URL') or os.getenv('JAWSDB_URL') or os.getenv('CLEARDB_DATABASE_URL')
if not db_url:
    mysql_user = os.getenv('MYSQL_USER')
    mysql_host = os.getenv('MYSQL_HOST')
    if mysql_user and mysql_host:
        db_url = (
            f"mysql+pymysql://{mysql_user}:"
            f"{os.getenv('MYSQL_PASSWORD', '')}@"
            f"{mysql_host}/"
            f"{os.getenv('MYSQL_DB', 'soiltrack_db')}"
        )
    else:
        # Fallback to lightweight SQLite database for cloud hosting if no MySQL env is set
        db_url = f"sqlite:///{os.path.join(app.root_path, 'soiltrack.db')}"

if db_url.startswith("mysql://"):
    db_url = db_url.replace("mysql://", "mysql+pymysql://", 1)
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

@app.errorhandler(500)
def internal_server_error(e):
    import traceback
    error_trace = traceback.format_exc()
    print("500 INTERNAL SERVER ERROR TRACEBACK:\n", error_trace)
    return f"<div style='font-family:monospace;padding:30px;'><h3>500 Internal Server Error</h3><pre>{error_trace}</pre></div>", 500

ROLE_LABELS = {
    'chemist': 'Chemist',
    'hod': 'HOD (Head of Dept.)',
    'admin': 'Admin',
    'field_officer': 'Field Officer',
    'staff': 'Chemist'
}

@app.context_processor
def inject_role_labels():
    return dict(role_labels=ROLE_LABELS)

# UNIT LABELS for display on calculation page / Excel headers
PARAMETER_UNITS = {
    'nitrogen': 'kg/ha',
    'phosphorus': 'kg/ha',
    'potassium': 'kg/ha',
    'organic_carbon': '%',
    'ec': 'mS/cm',
    'sulphur': 'ppm',
    'boron': 'ppm',
    'zinc': 'ppm',
    'iron': 'ppm',
    'manganese': 'ppm',
    'copper': 'ppm',
}

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fullname = db.Column(db.String(100))
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='chemist')
    # Task 3: Login lockout tracking
    failed_attempts = db.Column(db.Integer, default=0)
    last_failed_at = db.Column(db.DateTime, nullable=True)

class Sample(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sample_id = db.Column(db.String(50), unique=True, nullable=False)
    village = db.Column(db.String(100), nullable=False)
    taluka = db.Column(db.String(100))
    sample_type = db.Column(db.String(20))
    farmer_name = db.Column(db.String(100))
    collection_date = db.Column(db.String(20))

    # NEW: basic collection-time info
    phone_number = db.Column(db.String(20))
    address = db.Column(db.String(255))
    survey_number = db.Column(db.String(50))
    sample_source = db.Column(db.String(20))   # 'govt' or 'private'
    scheme = db.Column(db.String(150))         # only relevant when sample_source == 'govt'
    crop = db.Column(db.String(100))           # e.g. 'Paddy', 'Mango (1st year)', 'Chikku (5th year)', etc.
    test_package = db.Column(db.String(255))   # Selected rate card option(s) for private samples
    testing_fee = db.Column(db.Float, default=0.0) # Fee in ₹


    ph = db.Column(db.Float)
    ec = db.Column(db.Float)
    nitrogen = db.Column(db.Float)
    phosphorus = db.Column(db.Float)
    potassium = db.Column(db.Float)
    iron = db.Column(db.Float)
    manganese = db.Column(db.Float)
    copper = db.Column(db.Float)
    zinc = db.Column(db.Float)
    boron = db.Column(db.Float)
    organic_carbon = db.Column(db.Float)
    sulphur = db.Column(db.Float)
    # temperature / moisture REMOVED per request
    category = db.Column(db.String(20))
    notes = db.Column(db.String(200))
    observed_ec = db.Column(db.Float)
    ec_temperature = db.Column(db.Float)
    ec_comp_factor = db.Column(db.Float)
    n_burette_a = db.Column(db.Float)
    n_burette_b = db.Column(db.Float)
    abs_phosphorus = db.Column(db.Float)
    abs_potassium = db.Column(db.Float)
    abs_organic_carbon = db.Column(db.Float)
    abs_boron = db.Column(db.Float)
    abs_sulphur = db.Column(db.Float)
    abs_zinc = db.Column(db.Float)
    abs_iron = db.Column(db.Float)
    abs_manganese = db.Column(db.Float)
    abs_copper = db.Column(db.Float)

    # Analyzed by / Checked by / Approved by — for Soil Health Card
    analyzed_by = db.Column(db.String(100))
    checked_by = db.Column(db.String(100))
    approved_by = db.Column(db.String(100))

class DilutionFactor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    parameter = db.Column(db.String(50), unique=True, nullable=False)
    label = db.Column(db.String(100))
    factor = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(30))

# ── NEW: Lab Calculation tables ──────────────────────────────────────────
# These normalize the per-parameter results out of Sample so a full history
# of test results can be kept per sample, and a single fertility summary
# (score_ratio) can be stored per sample without recomputing it every time.

class TestResult(db.Model):
    """One row per chemistry parameter reading for a sample."""
    id = db.Column(db.Integer, primary_key=True)
    sample_id = db.Column(db.Integer, db.ForeignKey('sample.id'), nullable=False)
    parameter_type = db.Column(db.String(50), nullable=False)   # e.g. 'nitrogen', 'ph'
    calculated_value = db.Column(db.Float)
    category = db.Column(db.String(30))    # e.g. 'High' / 'Medium' / 'Low' / 'Sufficient' / 'Deficient' / 'Optimal'
    unit = db.Column(db.String(30))

    sample = db.relationship('Sample', backref=db.backref('test_results', lazy=True))

class LabCalculation(db.Model):
    """One row per sample — the overall fertility summary."""
    id = db.Column(db.Integer, primary_key=True)
    sample_id = db.Column(db.Integer, db.ForeignKey('sample.id'), unique=True, nullable=False)
    score_ratio = db.Column(db.Float)          # 0.0 - 1.0
    category = db.Column(db.String(20))        # 'Fertile' / 'Moderate' / 'Poor'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    sample = db.relationship('Sample', backref=db.backref('lab_calculation', uselist=False, lazy=True))

class LabCalculationTestResult(db.Model):
    """Many-to-many link between a LabCalculation and the TestResult rows that fed it."""
    id = db.Column(db.Integer, primary_key=True)
    lab_calculation_id = db.Column(db.Integer, db.ForeignKey('lab_calculation.id'), nullable=False)
    result_id = db.Column(db.Integer, db.ForeignKey('test_result.id'), nullable=False)

def seed_dilution_factors():
    defaults = [
        ('n_burette_b',     'Nitrogen Blank Reading (Burette B)', 0.0,    'ml'),
        ('phosphorus',     'Phosphorus (P)',                      41.65,  'kg/ha'),
        ('potassium',      'Potassium (K)',                        11.2,   'kg/ha'),
        ('organic_carbon', 'Organic Carbon (OC)',                 2.78,   '%'),
        ('boron',          'Boron (B)',                            5.36,   'ppm'),
        ('sulphur',        'Sulphur (S)',                          541.0,  'ppm'),
        ('zinc',           'Zinc (Zn)',                            18.8,   'ppm'),
        ('iron',           'Iron (Fe)',                            72.6,   'ppm'),
        ('manganese',      'Manganese (Mn)',                       39.87,  'ppm'),
        ('copper',         'Copper (Cu)',                          32.98,  'ppm'),
    ]
    for param, label, factor, unit in defaults:
        exists = DilutionFactor.query.filter_by(parameter=param).first()
        if not exists:
            db.session.add(DilutionFactor(parameter=param, label=label, factor=factor, unit=unit))
    db.session.commit()

# ── Auth Helpers ──
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to continue.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def staff_or_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to continue.', 'error')
            return redirect(url_for('login'))
        if session.get('role') not in ('chemist', 'hod', 'field_officer', 'admin', 'staff'):
            flash('You do not have permission to do that.', 'error')
            return redirect(url_for('all_samples'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to continue.', 'error')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Only Admin can do that.', 'error')
            return redirect(url_for('all_samples'))
        return f(*args, **kwargs)
    return decorated

EC_COMP_TABLE = {
    3.0:1.709,4.0:1.660,5.0:1.613,6.0:1.569,7.0:1.528,
    8.0:1.488,9.0:1.448,10.0:1.411,11.0:1.375,12.0:1.341,
    13.0:1.309,14.0:1.277,15.0:1.247,16.0:1.218,17.0:1.189,
    18.0:1.163,19.0:1.136,20.0:1.112,21.0:1.087,22.0:1.064,
    23.0:1.043,24.0:1.020,25.0:1.000,26.0:0.979,27.0:0.960,
    28.0:0.943,29.0:0.925,30.0:0.907,31.0:0.890,32.0:0.873,
    33.0:0.858,34.0:0.843,35.0:0.829,36.0:0.815,37.0:0.801,
    38.0:0.788,39.0:0.775,40.0:0.763,
}

def get_ec_comp_factor(temp):
    if not temp:
        return 1.0
    closest = min(EC_COMP_TABLE.keys(), key=lambda t: abs(t - temp))
    return EC_COMP_TABLE[closest]

def calculate_finals_from_raw(raw, factors):
    result = {}
    if raw.get('observed_ec') and raw.get('ec_temperature'):
        comp = get_ec_comp_factor(raw['ec_temperature'])
        result['ec'] = round(raw['observed_ec'] * comp, 4)
        result['ec_comp_factor'] = comp
    if raw.get('n_burette_a') is not None and raw.get('n_burette_b') is not None:
        result['nitrogen'] = round((raw['n_burette_a'] - raw['n_burette_b']) * 209.07, 2)
    abs_map = {
        'phosphorus':     'abs_phosphorus',
        'potassium':      'abs_potassium',
        'organic_carbon': 'abs_organic_carbon',
        'boron':          'abs_boron',
        'sulphur':        'abs_sulphur',
        'zinc':           'abs_zinc',
        'iron':           'abs_iron',
        'manganese':      'abs_manganese',
        'copper':         'abs_copper',
    }
    for param, abs_key in abs_map.items():
        if raw.get(abs_key) is not None:
            f = factors.get(param, 1.0)
            result[param] = round(raw[abs_key] * f, 4)
    return result

def get_factors_dict():
    factors = DilutionFactor.query.all()
    return {f.parameter: f.factor for f in factors}

TALUKA_CODES = {
    'ratnagiri': 'RT',
    'sangameshwar': 'SG',
    'sangameshar': 'SG',
    'lanja': 'LJ',
    'chiplun': 'CH',
    'rajapur': 'RJ',
    'guhagar': 'GH',
    'dapoli': 'DP',
    'khed': 'KH',
    'mandangad': 'MN'
}

def detect_taluka_code(taluka=None, address=None, village=None):
    text = f"{taluka or ''} {address or ''} {village or ''}".lower()
    for name, code in TALUKA_CODES.items():
        if name in text:
            return code
    return ''

def generate_sample_id(village, taluka=None, address=None):
    v_code = village[:3].upper() if len(village) >= 3 else village.upper().ljust(3, 'X')
    t_code = detect_taluka_code(taluka, address, village)

    if t_code:
        prefix = f"{t_code}-{v_code}"
    else:
        prefix = v_code

    cnt = Sample.query.filter(Sample.sample_id.like(f"{prefix}%")).count()
    seq_str = str(cnt + 1).zfill(2)
    return f"{prefix}{seq_str}"

def get_fertility_score(ph, nitrogen, phosphorus, potassium):
    """Returns (category, ratio) — same Fertile/Moderate/Poor logic as before,
    but now also exposes the raw score_ratio so it can be stored on LabCalculation."""
    score = 0
    total = 0
    if ph:
        total += 1
        if 6.5 <= ph <= 7.5:
            score += 1
    if nitrogen:
        total += 1
        if nitrogen >= 280:
            score += 1
    if phosphorus:
        total += 1
        if phosphorus >= 11:
            score += 1
    if potassium:
        total += 1
        if potassium >= 110:
            score += 1
    if total == 0:
        return "Poor", 0.0
    ratio = score / total
    if ratio >= 0.75:
        category = "Fertile"
    elif ratio >= 0.5:
        category = "Moderate"
    else:
        category = "Poor"
    return category, ratio

def get_category(ph, nitrogen, phosphorus, potassium):
    """Kept for backward compatibility with existing call sites — returns just the category."""
    category, _ratio = get_fertility_score(ph, nitrogen, phosphorus, potassium)
    return category

# ── NEW: per-parameter status thresholds ─────────────────────────────────
# Mirrors the exact thresholds used in soil_health_card.html so the
# TestResult.category values always match what the printed card shows.
def get_param_status(parameter_type, value):
    if value is None:
        return None

    if parameter_type == 'ph':
        if 6.5 <= value <= 7.5:
            return 'Optimal'
        elif 6.0 <= value < 6.5 or 7.5 < value <= 8.0:
            return 'Acceptable'
        else:
            return 'Needs Correction'

    if parameter_type == 'ec':
        if value <= 0.8:
            return 'Normal'
        elif value <= 1.6:
            return 'Slightly High'
        else:
            return 'High'

    if parameter_type == 'organic_carbon':
        if value >= 0.75:
            return 'High'
        elif value >= 0.5:
            return 'Medium'
        else:
            return 'Low'

    low_med_high = {
        'nitrogen':   (280, 560),
        'phosphorus': (11, 22),
        'potassium':  (110, 280),
    }
    if parameter_type in low_med_high:
        low, high = low_med_high[parameter_type]
        if value >= high:
            return 'High'
        elif value >= low:
            return 'Medium'
        else:
            return 'Low'

    sufficiency = {
        'sulphur':   10,
        'zinc':      0.6,
        'boron':     0.5,
        'iron':      4.5,
        'manganese': 2.0,
        'copper':    0.2,
    }
    if parameter_type in sufficiency:
        return 'Sufficient' if value >= sufficiency[parameter_type] else 'Deficient'

    return None

# ── NEW: fertiliser recommendation builder ───────────────────────────────
# Ports the exact rec-box logic from soil_health_card.html into Python so
# it can be stored / reused instead of only existing as Jinja conditionals.
def build_recommendation(sample):
    recs = []
    if sample.nitrogen is not None and sample.nitrogen < 280:
        recs.append('Apply Urea or DAP to increase Nitrogen')
    if sample.phosphorus is not None and sample.phosphorus < 11:
        recs.append('Apply SSP or DAP to increase Phosphorus')
    if sample.potassium is not None and sample.potassium < 110:
        recs.append('Apply MOP (Muriate of Potash) for Potassium')
    if sample.organic_carbon is not None and sample.organic_carbon < 0.75:
        recs.append('Add compost or FYM to improve Organic Carbon')
    if sample.zinc is not None and sample.zinc < 0.6:
        recs.append('Apply Zinc Sulphate @ 25 kg/ha')
    if sample.boron is not None and sample.boron < 0.5:
        recs.append('Apply Borax @ 10 kg/ha for Boron deficiency')
    if sample.sulphur is not None and sample.sulphur < 10:
        recs.append('Apply Elemental Sulphur or Gypsum')
    if sample.ph is not None and sample.ph < 6.0:
        recs.append('Apply Agricultural Lime to raise pH')
    if sample.ph is not None and sample.ph > 8.0:
        recs.append('Apply Gypsum or Sulphur to lower pH')
    if sample.category == 'Fertile':
        recs.append('Soil is healthy — maintain current practices')
        recs.append('Re-test after 6 months for best results')
    return recs

# ── NEW: sync_lab_tables ──────────────────────────────────────────────────
# Populates TestResult (per-parameter rows) and LabCalculation (per-sample
# fertility summary), and links them via LabCalculationTestResult.
# Call this AFTER sample.category has been set and the session has the
# latest chemistry values on `sample` (doesn't commit sample itself —
# caller is expected to commit once at the end).
def sync_lab_tables(sample):
    # 1. Fertility summary — one LabCalculation row per sample (create or update)
    category, ratio = get_fertility_score(sample.ph, sample.nitrogen, sample.phosphorus, sample.potassium)
    sample.category = category  # keep Sample.category in sync (used by dashboard/exports/health card)

    lab_calc = LabCalculation.query.filter_by(sample_id=sample.id).first()
    if not lab_calc:
        lab_calc = LabCalculation(sample_id=sample.id)
        db.session.add(lab_calc)
    lab_calc.score_ratio = ratio
    lab_calc.category = category
    lab_calc.updated_at = datetime.utcnow()
    db.session.flush()  # ensure lab_calc.id is available without a full commit

    # 2. Per-parameter TestResult rows
    all_params = {
        'ph': sample.ph,
        'ec': sample.ec,
        'organic_carbon': sample.organic_carbon,
        'nitrogen': sample.nitrogen,
        'phosphorus': sample.phosphorus,
        'potassium': sample.potassium,
        'sulphur': sample.sulphur,
        'zinc': sample.zinc,
        'boron': sample.boron,
        'iron': sample.iron,
        'manganese': sample.manganese,
        'copper': sample.copper,
    }

    for param, value in all_params.items():
        if value is None:
            continue
        status = get_param_status(param, value)
        unit = PARAMETER_UNITS.get(param, '')

        result = TestResult.query.filter_by(sample_id=sample.id, parameter_type=param).first()
        if not result:
            result = TestResult(sample_id=sample.id, parameter_type=param)
            db.session.add(result)
        result.calculated_value = value
        result.category = status
        result.unit = unit
        db.session.flush()  # ensure result.id is available

        # 3. Link into the junction table if not already linked
        link = LabCalculationTestResult.query.filter_by(
            lab_calculation_id=lab_calc.id, result_id=result.id
        ).first()
        if not link:
            db.session.add(LabCalculationTestResult(
                lab_calculation_id=lab_calc.id, result_id=result.id
            ))

# ── Auth Routes ──
@app.route('/login', methods=['GET'])
@app.route('/login/admin', methods=['GET'])
@app.route('/admin', methods=['GET'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    show_admin = (request.path in ['/admin', '/login/admin']) or (request.args.get('role') == 'admin') or (request.args.get('admin') == 'true')
    return render_template('login.html', error=None, show_admin=show_admin)

@app.route('/login', methods=['POST'])
@app.route('/login/admin', methods=['POST'])
@app.route('/admin', methods=['POST'])
def login_post():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role_selected = request.form.get('role', 'chemist')
    show_admin = (request.path in ['/admin', '/login/admin']) or (request.form.get('show_admin') == 'true') or (role_selected == 'admin')

    user = User.query.filter_by(username=username).first()

    # Task 3: Check lockout status (5 failed attempts within 15 minutes)
    if user:
        if (user.failed_attempts or 0) >= 5 and user.last_failed_at:
            elapsed = (datetime.now() - user.last_failed_at).total_seconds()
            if elapsed < 900:  # 15 minutes
                remaining_mins = max(1, int((900 - elapsed) // 60 + 1))
                return render_template('login.html', error=f'Too many failed attempts. Account locked. Please try again in {remaining_mins} minute(s).', show_admin=show_admin)
            else:
                user.failed_attempts = 0
                user.last_failed_at = None
                db.session.commit()

    if not user or not check_password_hash(user.password, password):
        if user:
            user.failed_attempts = (user.failed_attempts or 0) + 1
            user.last_failed_at = datetime.now()
            db.session.commit()
            remaining = max(0, 5 - user.failed_attempts)
            if remaining > 0:
                err_msg = f'Invalid username or password. ({remaining} attempt(s) remaining before account lockout)'
            else:
                err_msg = 'Too many failed attempts. Account locked for 15 minutes.'
            return render_template('login.html', error=err_msg, show_admin=show_admin)
        return render_template('login.html', error='Invalid username or password.', show_admin=show_admin)

    # Task 2: Validate selected role matches user.role
    user_role = user.role or 'chemist'
    if user_role != role_selected and not (user_role == 'staff' and role_selected == 'chemist'):
        user_label = ROLE_LABELS.get(user_role, user_role.title())
        selected_label = ROLE_LABELS.get(role_selected, role_selected.title())
        return render_template('login.html', error=f'This account is registered as "{user_label}", not "{selected_label}". Please select the correct role.', show_admin=show_admin)

    # Reset failed login counter on success
    if user.failed_attempts or user.last_failed_at:
        user.failed_attempts = 0
        user.last_failed_at = None
        db.session.commit()

    # Task 5: Set permanent session (30-min lifetime)
    session.permanent = True
    session['user_id'] = user.id
    session['username'] = user.username
    session['fullname'] = user.fullname
    session['role'] = user.role
    flash(f'Welcome back, {user.fullname or user.username}!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/register', methods=['GET'])
def register():
    list(get_flashed_messages())  # clear any stray flash messages
    admin_count = User.query.filter_by(role='admin').count()
    can_create_admin = admin_count < 2
    return render_template('register.html', error=None, can_create_admin=can_create_admin)

@app.route('/register', methods=['POST'])
def register_post():
    fullname = request.form.get('fullname', '').strip()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    confirm_password = request.form.get('confirm_password', '')
    role = request.form.get('role', 'chemist')
    if role not in ('chemist', 'hod', 'field_officer', 'admin'):
        role = 'chemist'

    admin_count = User.query.filter_by(role='admin').count()
    can_create_admin = admin_count < 2

    if role == 'admin' and not can_create_admin:
        return render_template(
            'register.html',
            error='Maximum limit of 2 Admin accounts has been reached for this system. Please register as Chemist or Field Officer.',
            can_create_admin=can_create_admin
        )

    if not username or not password:
        return render_template('register.html', error='Username and password are required.', can_create_admin=can_create_admin)

    if password != confirm_password:
        return render_template('register.html', error='Passwords do not match. Please try again.', can_create_admin=can_create_admin)

    existing = User.query.filter_by(username=username).first()
    if existing:
        return render_template('register.html', error='Username already taken. Please choose another.', can_create_admin=can_create_admin)

    new_user = User(
        fullname=fullname,
        username=username,
        password=generate_password_hash(password),
        role=role
    )
    db.session.add(new_user)
    db.session.commit()
    flash('Account created successfully! Please login.', 'success')
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['POST'])
def forgot_password_post():
    username = request.form.get('username', '').strip()
    fullname = request.form.get('fullname', '').strip()
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    if not username or not fullname or not new_password:
        flash('Username, Full Name, and new password are required.', 'error')
        return redirect(url_for('login'))

    if new_password != confirm_password:
        flash('New passwords do not match. Please try again.', 'error')
        return redirect(url_for('login'))

    user = User.query.filter_by(username=username).first()
    if not user:
        flash('User account not found with that username.', 'error')
        return redirect(url_for('login'))

    # Verify registered Full Name matches (case-insensitive)
    if user.fullname and user.fullname.strip().lower() != fullname.lower():
        flash('Full Name verification failed. Please enter your registered Full Name.', 'error')
        return redirect(url_for('login'))

    user.password = generate_password_hash(new_password)
    user.failed_attempts = 0
    user.last_failed_at = None
    db.session.commit()

    flash('🎉 Password reset successfully! You can now log in with your new password.', 'success')
    return redirect(url_for('login'))

def seed_admin_account():
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        db.session.add(User(
            fullname='Administrator',
            username='admin',
            password=generate_password_hash('admin123'),
            role='admin'
        ))
        db.session.commit()

def init_db_tables():
    with app.app_context():
        try:
            db.create_all()
            seed_dilution_factors()
            seed_admin_account()
        except Exception as e:
            print("Database initialization notice:", e)

init_db_tables()

# ── Dashboard ──
@app.route('/')
@login_required
def dashboard():
    total = Sample.query.count()
    govt_count = Sample.query.filter(
        or_(Sample.sample_source == 'govt', Sample.sample_type == 'Government')
    ).count()
    pvt_count = Sample.query.filter(
        or_(Sample.sample_source == 'private', Sample.sample_type == 'Private', Sample.sample_source == None)
    ).count()
    recent = Sample.query.order_by(Sample.id.desc()).limit(8).all()

    return render_template('dashboard.html',
        total=total,
        govt_count=govt_count,
        pvt_count=pvt_count,
        recent=recent)

# ── Yearly Records & Annual Report Archive ──
@app.route('/yearly-records')
@login_required
def yearly_records():
    selected_year = request.args.get('year', '')
    
    # Extract all distinct collection years from database
    raw_dates = [s.collection_date for s in Sample.query.all() if s.collection_date]
    years = sorted(list(set(raw_dates)), reverse=True)
    if not years:
        years = ['2025-2026']

    query = Sample.query
    if selected_year:
        query = query.filter(Sample.collection_date == selected_year)
        
    samples = query.all()
    
    # Annual statistics calculation
    total_count = len(samples)
    govt_count = sum(1 for s in samples if (s.sample_source == 'govt' or s.sample_type == 'Government'))
    pvt_count = sum(1 for s in samples if (s.sample_source == 'private' or s.sample_type == 'Private' or not s.sample_source))
    total_revenue = sum((s.testing_fee or 0.0) for s in samples if (s.sample_source == 'private' or s.sample_type == 'Private' or not s.sample_source))
    
    # Parameter Averages
    ph_vals = [s.ph for s in samples if s.ph is not None]
    ec_vals = [s.ec for s in samples if s.ec is not None]
    oc_vals = [s.organic_carbon for s in samples if s.organic_carbon is not None]
    n_vals = [s.nitrogen for s in samples if s.nitrogen is not None]
    p_vals = [s.phosphorus for s in samples if s.phosphorus is not None]
    k_vals = [s.potassium for s in samples if s.potassium is not None]
    
    avg_ph = round(sum(ph_vals)/len(ph_vals), 2) if ph_vals else None
    avg_ec = round(sum(ec_vals)/len(ec_vals), 3) if ec_vals else None
    avg_oc = round(sum(oc_vals)/len(oc_vals), 2) if oc_vals else None
    avg_n = round(sum(n_vals)/len(n_vals), 1) if n_vals else None
    avg_p = round(sum(p_vals)/len(p_vals), 1) if p_vals else None
    avg_k = round(sum(k_vals)/len(k_vals), 1) if k_vals else None
    
    # Village breakdown for this year
    village_stats = {}
    for s in samples:
        v = s.village or 'Unknown'
        if v not in village_stats:
            village_stats[v] = {'total': 0, 'govt': 0, 'pvt': 0}
        village_stats[v]['total'] += 1
        if s.sample_source == 'govt' or s.sample_type == 'Government':
            village_stats[v]['govt'] += 1
        else:
            village_stats[v]['pvt'] += 1

    return render_template('yearly_records.html',
        years=years,
        selected_year=selected_year,
        samples=samples,
        total_count=total_count,
        govt_count=govt_count,
        pvt_count=pvt_count,
        total_revenue=total_revenue,
        avg_ph=avg_ph,
        avg_ec=avg_ec,
        avg_oc=avg_oc,
        avg_n=avg_n,
        avg_p=avg_p,
        avg_k=avg_k,
        village_stats=village_stats)

# ── Samples ──
@app.route('/samples')
@login_required
def all_samples():
    samples = Sample.query.all()
    return render_template('samples.html', samples=samples)

@app.route('/add')
@staff_or_admin_required
def add_sample():
    return render_template('add_sample.html')

@app.route('/add', methods=['POST'])
@staff_or_admin_required
def save_sample():
    """
    Collection-time registration ONLY.
    No chemistry parameters here anymore — those are entered later
    via the Lab Calculation wizard once the sample is actually tested.
    """
    village = request.form.get('village')
    sample_source = request.form.get('sample_source', 'private')
    scheme = request.form.get('scheme') if sample_source == 'govt' else None
    crop = request.form.get('crop')
    test_package = request.form.get('test_package') if sample_source == 'private' else None
    
    fee_str = request.form.get('testing_fee', '0')
    try:
        testing_fee = float(fee_str) if (sample_source == 'private' and fee_str) else 0.0
    except ValueError:
        testing_fee = 0.0

    new_sample = Sample(
        sample_id       = generate_sample_id(village, taluka=request.form.get('taluka'), address=request.form.get('address')),
        village         = village,
        sample_type     = request.form.get('sample_type'),
        farmer_name     = request.form.get('farmer_name'),
        collection_date = request.form.get('collection_date'),
        phone_number    = request.form.get('phone_number'),
        address         = request.form.get('address'),
        survey_number   = request.form.get('survey_number'),
        sample_source   = sample_source,
        scheme          = scheme,
        crop            = crop,
        test_package    = test_package,
        testing_fee     = testing_fee,
        notes           = request.form.get('notes'),
        category        = None,  # unknown until lab calculation is done
    )
    db.session.add(new_sample)
    db.session.commit()
    flash(f'Sample {new_sample.sample_id} registered. Parameters can be added later via Lab Calculation.', 'success')
    return redirect(url_for('dashboard'))

# ── Bulk Sample Import ──
@app.route('/download-import-template')
@staff_or_admin_required
def download_import_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['village', 'farmer_name', 'phone_number', 'address', 'survey_number', 'sample_source', 'scheme', 'crop', 'testing_fee', 'collection_date'])
    writer.writerow(['Ratnagiri', 'Ramesh Patil', '9876543210', 'Plot 12, Main Road', 'S-101', 'govt', 'PM-KISAN', 'Rice', '0', '2026-07-30'])
    writer.writerow(['Chiplun', 'Suresh Joshi', '9123456789', 'Near Gram Panchayat', 'S-102', 'private', '', 'Wheat', '250', '2026-07-30'])
    
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=SoilTrack_Bulk_Import_Template.csv'
    return response

@app.route('/bulk-add', methods=['GET'])
@staff_or_admin_required
def bulk_add():
    return render_template('bulk_add.html')

TEMP_IMPORT_DIR = os.path.join(app.root_path, 'temp_imports')
os.makedirs(TEMP_IMPORT_DIR, exist_ok=True)

TARGET_FIELDS = [
    ('sample_id', 'Sample ID (For updating existing sample)', False),
    ('village', 'Village Name', True),
    ('taluka', 'Taluka / Block', False),
    ('farmer_name', 'Farmer Full Name', False),
    ('phone_number', 'Mobile / Phone Number', False),
    ('address', 'Farmer Address', False),
    ('survey_number', 'Survey / Gat / Plot No', False),
    ('sample_source', 'Sample Source (Govt / Private)', False),
    ('scheme', 'Government Scheme', False),
    ('crop', 'Crop / Sample Type', False),
    ('testing_fee', 'Testing Fee', False),
    ('collection_date', 'Collection Date', False),
    ('ph', 'pH Value', False),
    ('ec', 'EC Value', False),
    ('organic_carbon', 'Organic Carbon (%)', False),
    ('nitrogen', 'Nitrogen (kg/ha)', False),
    ('phosphorus', 'Phosphorus (kg/ha)', False),
    ('potassium', 'Potassium (kg/ha)', False),
    ('sulphur', 'Sulphur (ppm)', False),
    ('zinc', 'Zinc (ppm)', False),
    ('boron', 'Boron (ppm)', False),
    ('iron', 'Iron (ppm)', False),
    ('manganese', 'Manganese (ppm)', False),
    ('copper', 'Copper (ppm)', False),
]

def auto_detect_header(target_key, raw_headers):
    keywords = {
        'sample_id': ['sample id', 'sample_id', 'test id', 'lab no', 'sample no', 'code'],
        'village': ['village', 'gaon', 'gram', 'city', 'place', 'location'],
        'taluka': ['taluka', 'taluk', 'block', 'tahsil', 'tehsil'],
        'farmer_name': ['farmer name', 'farmer', 'kisan', 'owner', 'holder'],
        'phone_number': ['phone number', 'phone', 'mobile', 'contact', 'cell', 'tel'],
        'address': ['farmer address', 'address', 'pata', 'street'],
        'survey_number': ['survey number', 'survey', 'gut', 'gat', 'khasra', 'plot'],
        'sample_source': ['source', 'category', 'type', 'govt'],
        'scheme': ['scheme', 'yojana', 'govt', 'government'],
        'crop': ['crop', 'piq', 'plant', 'sample_type'],
        'testing_fee': ['fee', 'cost', 'amount', 'price', 'charge'],
        'collection_date': ['date', 'tikh', 'time'],
        'ph': ['ph', 'ph value', 'soil ph', 'ph_val'],
        'ec': ['ec', 'conductivity'],
        'organic_carbon': ['oc', 'organic carbon', 'carbon'],
        'nitrogen': ['nitrogen', 'n_val', 'n '],
        'phosphorus': ['phosphorus', 'p_val', 'p '],
        'potassium': ['potassium', 'k_val', 'k '],
        'sulphur': ['sulphur', 'sulfur', 's_val'],
        'zinc': ['zinc', 'zn'],
        'boron': ['boron', 'b_val'],
        'iron': ['iron', 'fe'],
        'manganese': ['manganese', 'mn'],
        'copper': ['copper', 'cu']
    }

    for h in raw_headers:
        h_clean = h.lower().replace('_', ' ').replace('-', ' ').strip()

        # SPECIAL SAFETY FILTER FOR pH: Do not match 'phone number' as 'pH'
        if target_key == 'ph':
            if 'phone' in h_clean:
                continue
            words = h_clean.split()
            if 'ph' in words or h_clean in ['ph', 'ph value', 'soil ph', 'ph_val'] or h_clean.startswith('ph ') or h_clean.endswith(' ph'):
                return h
            continue

        target_kws = keywords.get(target_key, [target_key])
        for kw in target_kws:
            if kw in h_clean:
                return h
    return ''

@app.route('/bulk-add', methods=['POST'])
@staff_or_admin_required
def bulk_add_post():
    uploaded_file = request.files.get('file')
    if not uploaded_file or not uploaded_file.filename:
        flash('Please select an Excel (.xlsx) or CSV (.csv) file to upload.', 'error')
        return redirect(url_for('bulk_add'))

    filename = uploaded_file.filename
    lower_filename = filename.lower()
    if not (lower_filename.endswith('.csv') or lower_filename.endswith('.xlsx') or lower_filename.endswith('.xls')):
        flash('Invalid file format. Please upload a .csv or .xlsx file.', 'error')
        return redirect(url_for('bulk_add'))

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    temp_filename = f"temp_{timestamp}_{filename}"
    temp_filepath = os.path.join(TEMP_IMPORT_DIR, temp_filename)
    uploaded_file.save(temp_filepath)

    raw_headers = []
    preview_rows = []
    total_rows = 0
    sheet_names = ['Sheet 1']
    selected_sheet = request.form.get('selected_sheet', '')

    try:
        if lower_filename.endswith('.csv'):
            with open(temp_filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.reader(f)
                header_row = next(reader, None)
                if header_row:
                    raw_headers = [str(h).strip() for h in header_row if str(h).strip()]
                    for i, row in enumerate(reader):
                        if any(row):
                            total_rows += 1
                            if i < 3:
                                row_dict = {}
                                for idx, cell_val in enumerate(row):
                                    if idx < len(raw_headers):
                                        row_dict[raw_headers[idx]] = str(cell_val).strip()
                                preview_rows.append(row_dict)

        elif lower_filename.endswith('.xlsx') or lower_filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(temp_filepath, data_only=True)
            sheet_names = wb.sheetnames
            if not selected_sheet or selected_sheet not in sheet_names:
                # Intelligently auto-select data sheet like 'cumulative' or 'data' or first sheet
                selected_sheet = sheet_names[0]
                for s in sheet_names:
                    if 'cumul' in s.lower() or 'data' in s.lower() or 'sample' in s.lower():
                        selected_sheet = s
                        break
            sheet = wb[selected_sheet]

            header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            raw_headers = [str(c).strip() for c in header_cells if c is not None and str(c).strip()]

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if not any(row):
                    continue
                total_rows += 1
                if i < 3:
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(raw_headers):
                            row_dict[raw_headers[idx]] = str(val or '').strip()
                    preview_rows.append(row_dict)

    except Exception as e:
        if os.path.exists(temp_filepath):
            os.remove(temp_filepath)
        flash(f'Error reading Excel file: {str(e)}', 'error')
        return redirect(url_for('bulk_add'))

    if not raw_headers or total_rows == 0:
        if os.path.exists(temp_filepath):
            os.remove(temp_filepath)
        flash('The uploaded Excel/CSV file is empty or has no header row.', 'error')
        return redirect(url_for('bulk_add'))

    auto_mappings = {field_key: auto_detect_header(field_key, raw_headers) for field_key, _, _ in TARGET_FIELDS}

    return render_template(
        'bulk_map.html',
        temp_filename=temp_filename,
        original_filename=filename,
        raw_headers=raw_headers,
        target_fields=TARGET_FIELDS,
        auto_mappings=auto_mappings,
        preview_rows=preview_rows,
        total_rows=total_rows,
        sheet_names=sheet_names,
        selected_sheet=selected_sheet
    )

@app.route('/bulk-switch-sheet', methods=['POST'])
@staff_or_admin_required
def bulk_switch_sheet_post():
    temp_filename = request.form.get('temp_filename')
    selected_sheet = request.form.get('selected_sheet', '')
    if not temp_filename:
        flash('Session expired. Please upload file again.', 'error')
        return redirect(url_for('bulk_add'))

    temp_filepath = os.path.join(TEMP_IMPORT_DIR, temp_filename)
    if not os.path.exists(temp_filepath):
        flash('Temporary import file missing.', 'error')
        return redirect(url_for('bulk_add'))

    import openpyxl
    wb = openpyxl.load_workbook(temp_filepath, data_only=True)
    sheet_names = wb.sheetnames
    if not selected_sheet or selected_sheet not in sheet_names:
        selected_sheet = sheet_names[0]
    sheet = wb[selected_sheet]

    header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    raw_headers = [str(c).strip() for c in header_cells if c is not None and str(c).strip()]

    preview_rows = []
    total_rows = 0
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not any(row):
            continue
        total_rows += 1
        if i < 3:
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(raw_headers):
                    row_dict[raw_headers[idx]] = str(val or '').strip()
            preview_rows.append(row_dict)

    auto_mappings = {field_key: auto_detect_header(field_key, raw_headers) for field_key, _, _ in TARGET_FIELDS}

    return render_template(
        'bulk_map.html',
        temp_filename=temp_filename,
        original_filename=temp_filename.split('_', 2)[-1] if '_' in temp_filename else temp_filename,
        raw_headers=raw_headers,
        target_fields=TARGET_FIELDS,
        auto_mappings=auto_mappings,
        preview_rows=preview_rows,
        total_rows=total_rows,
        sheet_names=sheet_names,
        selected_sheet=selected_sheet
    )

@app.route('/bulk-process', methods=['POST'])
@staff_or_admin_required
def bulk_process_post():
    temp_filename = request.form.get('temp_filename')
    if not temp_filename:
        flash('Session expired or file missing. Please re-upload your Excel file.', 'error')
        return redirect(url_for('bulk_add'))

    temp_filepath = os.path.join(TEMP_IMPORT_DIR, temp_filename)
    if not os.path.exists(temp_filepath):
        flash('Temporary import file not found. Please upload again.', 'error')
        return redirect(url_for('bulk_add'))

    # User Mappings
    col_sample_id = request.form.get('map_sample_id', '')
    col_village = request.form.get('map_village', '')
    col_taluka = request.form.get('map_taluka', '')
    col_farmer = request.form.get('map_farmer_name', '')
    col_phone = request.form.get('map_phone_number', '')
    col_address = request.form.get('map_address', '')
    col_survey = request.form.get('map_survey_number', '')
    col_source = request.form.get('map_sample_source', '')
    col_scheme = request.form.get('map_scheme', '')
    col_crop = request.form.get('map_crop', '')
    col_fee = request.form.get('map_testing_fee', '')
    col_date = request.form.get('map_collection_date', '')

    # Test Parameter Mappings
    col_ph = request.form.get('map_ph', '')
    col_ec = request.form.get('map_ec', '')
    col_oc = request.form.get('map_organic_carbon', '')
    col_n = request.form.get('map_nitrogen', '')
    col_p = request.form.get('map_phosphorus', '')
    col_k = request.form.get('map_potassium', '')
    col_s = request.form.get('map_sulphur', '')
    col_zn = request.form.get('map_zinc', '')
    col_b = request.form.get('map_boron', '')
    col_fe = request.form.get('map_iron', '')
    col_mn = request.form.get('map_manganese', '')
    col_cu = request.form.get('map_copper', '')

    # Default Assignments & Conflict Resolution
    default_source = request.form.get('default_sample_source', 'govt')
    default_scheme = request.form.get('default_scheme', '').strip()
    fill_blanks_only = request.form.get('fill_blanks_only') == 'on'
    selected_sheet = request.form.get('selected_sheet', '')

    if not col_village and not col_sample_id:
        flash('Please select which column maps to "Village Name" or "Sample ID".', 'error')
        return redirect(url_for('bulk_add'))

    rows_data = []
    lower_filename = temp_filename.lower()

    try:
        if lower_filename.endswith('.csv'):
            with open(temp_filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.DictReader(f)
                for r in reader:
                    rows_data.append({str(k).strip(): str(v).strip() for k, v in r.items() if k})

        elif lower_filename.endswith('.xlsx') or lower_filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(temp_filepath, data_only=True)
            sheet = wb[selected_sheet] if selected_sheet and selected_sheet in wb.sheetnames else wb.active
            headers = [str(c.value or '').strip() for c in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = str(val or '').strip()
                rows_data.append(row_dict)

    except Exception as e:
        flash(f'Error processing file rows: {str(e)}', 'error')
        return redirect(url_for('bulk_add'))
    finally:
        if os.path.exists(temp_filepath):
            os.remove(temp_filepath)

    if not rows_data:
        flash('No sample data rows found to process.', 'error')
        return redirect(url_for('bulk_add'))

    imported_count = 0
    updated_count = 0
    village_counts = {}
    first_sample_id = None
    last_sample_id = None
    today_str = datetime.now().strftime('%Y-%m-%d')

    # Extract village fallback from filename if available (e.g., Jambhulwadi)
    filename_village = 'Jambhulwadi'
    clean_filename = temp_filename.split('_', 2)[-1] if '_' in temp_filename else temp_filename
    for word in clean_filename.replace('.xlsx', '').replace('.xls', '').replace('.csv', '').split('_'):
        if word.lower() not in ['soil', 'data', 'analysis', 'final', 'uploading', 'temp', 'import', 'sheet']:
            filename_village = word.strip().capitalize()
            break

    for r in rows_data:
        raw_mapped_id = (r.get(col_sample_id, '') if col_sample_id else '').strip()
        # Only use mapped_sample_id if it looks like an actual sample code (e.g. JAM01, RAT01), not plain row numbers like 1, 2, 3
        mapped_sample_id = raw_mapped_id if (raw_mapped_id and not raw_mapped_id.isdigit() and len(raw_mapped_id) >= 3) else ''

        village = (r.get(col_village, '') if col_village else '').strip()
        taluka = (r.get(col_taluka, '') if col_taluka else '').strip()
        
        # Check if existing sample exists by Sample ID OR by (Village + Farmer Name + Survey No)
        existing_sample = None
        if mapped_sample_id:
            existing_sample = Sample.query.filter_by(sample_id=mapped_sample_id).first()
        
        farmer_name = (r.get(col_farmer, '') if col_farmer else '').strip()
        phone_number = (r.get(col_phone, '') if col_phone else '').strip()
        address = (r.get(col_address, '') if col_address else '').strip()
        survey_number = (r.get(col_survey, '') if col_survey else '').strip()

        if not existing_sample and survey_number and farmer_name and village:
            existing_sample = Sample.query.filter_by(village=village, farmer_name=farmer_name, survey_number=survey_number).first()

        if not village:
            village = existing_sample.village if (existing_sample and existing_sample.village) else filename_village

        raw_src = (r.get(col_source, '') if col_source else default_source).lower()
        sample_source = 'govt' if 'gov' in raw_src else ('private' if 'pvt' in raw_src or 'priv' in raw_src else default_source)
        
        scheme_val = (r.get(col_scheme, '') if col_scheme else default_scheme).strip()
        scheme = scheme_val if sample_source == 'govt' else None
        crop = (r.get(col_crop, '') if col_crop else '').strip()
        collection_date = (r.get(col_date, '') if col_date else '').strip() or today_str

        fee_val = (r.get(col_fee, '0') if col_fee else '0').strip()
        try:
            testing_fee = float(fee_val) if sample_source == 'private' else 0.0
        except ValueError:
            testing_fee = 0.0

        # Helper to parse float parameters safely
        def parse_float(key_col):
            if not key_col: return None
            val_str = r.get(key_col, '').strip()
            if not val_str: return None
            try:
                return float(val_str)
            except ValueError:
                return None

        val_ph = parse_float(col_ph)
        val_ec = parse_float(col_ec)
        val_oc = parse_float(col_oc)
        val_n = parse_float(col_n)
        val_p = parse_float(col_p)
        val_k = parse_float(col_k)
        val_s = parse_float(col_s)
        val_zn = parse_float(col_zn)
        val_b = parse_float(col_b)
        val_fe = parse_float(col_fe)
        val_mn = parse_float(col_mn)
        val_cu = parse_float(col_cu)

        # IF SAMPLE EXISTS: UPDATE / REPLACE EXISTING SAMPLE RECORD
        if existing_sample:
            if farmer_name: existing_sample.farmer_name = farmer_name
            if phone_number: existing_sample.phone_number = phone_number
            if address: existing_sample.address = address
            if crop: existing_sample.crop = crop
            if scheme: existing_sample.scheme = scheme
            if testing_fee > 0: existing_sample.testing_fee = testing_fee
            
            # Update test parameters if present in re-uploaded file
            if val_ph is not None: existing_sample.ph = val_ph
            if val_ec is not None: existing_sample.ec = val_ec
            if val_oc is not None: existing_sample.organic_carbon = val_oc
            if val_n is not None: existing_sample.nitrogen = val_n
            if val_p is not None: existing_sample.phosphorus = val_p
            if val_k is not None: existing_sample.potassium = val_k
            if val_s is not None: existing_sample.sulphur = val_s
            if val_zn is not None: existing_sample.zinc = val_zn
            if val_b is not None: existing_sample.boron = val_b
            if val_fe is not None: existing_sample.iron = val_fe
            if val_mn is not None: existing_sample.manganese = val_mn
            if val_cu is not None: existing_sample.copper = val_cu
            
            updated_count += 1
            continue

        # IF NEW SAMPLE: CREATE NEW RECORD WITH AUTO-GENERATED ID (SG-PAL01, RT-PAL01, JAM01...)
        sample_id = generate_sample_id(village, taluka=taluka, address=address)

        if not first_sample_id:
            first_sample_id = sample_id
        last_sample_id = sample_id

        new_sample = Sample(
            sample_id=sample_id,
            village=village,
            farmer_name=farmer_name,
            phone_number=phone_number,
            address=address,
            survey_number=survey_number,
            sample_source=sample_source,
            scheme=scheme,
            crop=crop,
            testing_fee=testing_fee,
            collection_date=collection_date,
            ph=val_ph,
            ec=val_ec,
            organic_carbon=val_oc,
            nitrogen=val_n,
            phosphorus=val_p,
            potassium=val_k,
            sulphur=val_s,
            zinc=val_zn,
            boron=val_b,
            iron=val_fe,
            manganese=val_mn,
            copper=val_cu
        )
        db.session.add(new_sample)
        imported_count += 1

    db.session.commit()

    id_range = f"{first_sample_id} to {last_sample_id}" if first_sample_id != last_sample_id else (first_sample_id or 'N/A')
    msg = f'🎉 Bulk Import Completed! Imported {imported_count} new soil samples (IDs: {id_range})'
    if updated_count > 0:
        msg += f' and updated {updated_count} existing records.'
    flash(msg, 'success')
    return redirect(url_for('all_samples'))

@app.route('/sample/<int:id>')
@login_required
def sample_detail(id):
    sample = db.session.get(Sample, id)
    if not sample:
        return redirect(url_for('all_samples'))
    return render_template('sample_detail.html', sample=sample)

@app.route('/edit/<int:id>')
@staff_or_admin_required
def edit_sample(id):
    sample = db.session.get(Sample, id)
    return render_template('edit_sample.html', sample=sample)

@app.route('/edit/<int:id>', methods=['POST'])
@staff_or_admin_required
def update_sample(id):
    """
    Edits basic collection-time info only (village, farmer, phone,
    address, survey number, scheme, crop, test package, fee, notes).
    """
    sample = db.session.get(Sample, id)
    sample_source = request.form.get('sample_source', sample.sample_source or 'private')
    scheme = request.form.get('scheme') if sample_source == 'govt' else None
    crop = request.form.get('crop')
    test_package = request.form.get('test_package') if sample_source == 'private' else None
    
    fee_str = request.form.get('testing_fee', '0')
    try:
        testing_fee = float(fee_str) if (sample_source == 'private' and fee_str) else 0.0
    except ValueError:
        testing_fee = 0.0

    sample.village         = request.form.get('village')
    taluka_input           = request.form.get('taluka')
    address_input          = request.form.get('address')

    if hasattr(sample, 'taluka'):
        sample.taluka = taluka_input

    # Dynamic Region / Taluka Code & Sample ID Prefix Update
    t_code = detect_taluka_code(taluka_input, address_input, sample.village)
    v_code = sample.village[:3].upper() if (sample.village and len(sample.village) >= 3) else (sample.village.upper().ljust(3, 'X') if sample.village else 'SMP')
    
    curr_id = sample.sample_id or ''
    seq_match = re.search(r'\d+$', curr_id)
    seq_num = seq_match.group(0) if seq_match else '01'

    if t_code:
        sample.sample_id = f"{t_code}-{v_code}{seq_num}"
    else:
        sample.sample_id = f"{v_code}{seq_num}"

    sample.sample_type     = request.form.get('sample_type')
    sample.farmer_name     = request.form.get('farmer_name')
    sample.collection_date = request.form.get('collection_date')
    sample.phone_number    = request.form.get('phone_number')
    sample.address         = address_input
    sample.survey_number   = request.form.get('survey_number')
    sample.sample_source   = sample_source
    sample.scheme          = scheme
    sample.crop            = crop
    sample.test_package    = test_package
    sample.testing_fee     = testing_fee
    sample.notes           = request.form.get('notes')
    db.session.commit()
    return redirect(url_for('sample_detail', id=sample.id))

@app.route('/delete/<int:id>')
@staff_or_admin_required
def delete_sample(id):
    sample = db.session.get(Sample, id)
    if sample:
        db.session.delete(sample)
        db.session.commit()
    return redirect(url_for('all_samples'))

@app.route('/bulk-delete', methods=['POST'])
@staff_or_admin_required
def bulk_delete_samples():
    sample_ids = request.form.getlist('sample_ids')
    if sample_ids:
        if 'all' in sample_ids:
            num = Sample.query.delete()
            db.session.commit()
            flash(f'✅ All {num} samples have been deleted from the database.', 'success')
        else:
            ids = [int(i) for i in sample_ids if i.isdigit()]
            if ids:
                num = Sample.query.filter(Sample.id.in_(ids)).delete(synchronize_session=False)
                db.session.commit()
                flash(f'✅ Successfully deleted {num} selected samples.', 'success')
    else:
        flash('No samples selected for deletion.', 'error')
    return redirect(url_for('all_samples'))

# ── Export ──
def _export_headers():
    return [
        'Sample ID', 'Village', 'Type', 'Farmer', 'Phone Number', 'Address',
        'Survey Number', 'Sample Source', 'Scheme', 'Date',
        'pH', f"EC ({PARAMETER_UNITS['ec']})", f"OC ({PARAMETER_UNITS['organic_carbon']})",
        f"Nitrogen ({PARAMETER_UNITS['nitrogen']})",
        f"Phosphorus ({PARAMETER_UNITS['phosphorus']})",
        f"Potassium ({PARAMETER_UNITS['potassium']})",
        f"Iron ({PARAMETER_UNITS['iron']})",
        f"Manganese ({PARAMETER_UNITS['manganese']})",
        f"Copper ({PARAMETER_UNITS['copper']})",
        f"Zinc ({PARAMETER_UNITS['zinc']})",
        f"Boron ({PARAMETER_UNITS['boron']})",
        f"Sulphur ({PARAMETER_UNITS['sulphur']})",
        'Category', 'Notes',
        'Observed EC', 'EC Temp', 'EC Comp Factor',
        'N Burette A', 'N Burette B',
        'Abs P', 'Abs K', 'Abs OC', 'Abs B', 'Abs S',
        'Abs Zn', 'Abs Fe', 'Abs Mn', 'Abs Cu'
    ]
    # NOTE: Temperature / Moisture columns removed per request

def _export_row(s):
    return [
        s.sample_id, s.village, s.sample_type, s.farmer_name,
        s.phone_number, s.address, s.survey_number,
        s.sample_source, s.scheme, s.collection_date,
        s.ph, s.ec, s.organic_carbon,
        s.nitrogen, s.phosphorus, s.potassium,
        s.iron, s.manganese, s.copper, s.zinc, s.boron, s.sulphur,
        s.category, s.notes,
        s.observed_ec, s.ec_temperature, s.ec_comp_factor,
        s.n_burette_a, s.n_burette_b,
        s.abs_phosphorus, s.abs_potassium, s.abs_organic_carbon,
        s.abs_boron, s.abs_sulphur, s.abs_zinc,
        s.abs_iron, s.abs_manganese, s.abs_copper
    ]

@app.route('/export')
@login_required
def export_excel():
    samples = Sample.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_export_headers())
    for s in samples:
        writer.writerow(_export_row(s))
    output.seek(0)
    return Response(output, mimetype='text/csv',
                    headers={"Content-Disposition": "attachment;filename=soiltrack_samples.csv"})

@app.route('/export-selected')
@login_required
def export_selected():
    ids_param = request.args.get('ids', '')
    if not ids_param:
        return redirect(url_for('all_samples'))
    ids = [int(i) for i in ids_param.split(',') if i.isdigit()]
    samples = Sample.query.filter(Sample.id.in_(ids)).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_export_headers())
    for s in samples:
        writer.writerow(_export_row(s))
    output.seek(0)
    filename = f'soiltrack_selected_{len(samples)}_samples.csv'
    return Response(output, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment;filename={filename}'})

# ── Recalculate ──
@app.route('/recalculate')
@admin_required
def recalculate():
    samples = Sample.query.all()
    factors = get_factors_dict()
    count = 0
    for s in samples:
        raw = {
            'observed_ec': s.observed_ec, 'ec_temperature': s.ec_temperature,
            'n_burette_a': s.n_burette_a, 'n_burette_b': s.n_burette_b,
            'abs_phosphorus': s.abs_phosphorus, 'abs_potassium': s.abs_potassium,
            'abs_organic_carbon': s.abs_organic_carbon, 'abs_boron': s.abs_boron,
            'abs_sulphur': s.abs_sulphur, 'abs_zinc': s.abs_zinc,
            'abs_iron': s.abs_iron, 'abs_manganese': s.abs_manganese,
            'abs_copper': s.abs_copper,
        }
        calc = calculate_finals_from_raw(raw, factors)
        if calc.get('ec'):             s.ec = calc['ec']
        if calc.get('nitrogen'):       s.nitrogen = calc['nitrogen']
        if calc.get('phosphorus'):     s.phosphorus = calc['phosphorus']
        if calc.get('potassium'):      s.potassium = calc['potassium']
        if calc.get('organic_carbon'): s.organic_carbon = calc['organic_carbon']
        if calc.get('boron'):          s.boron = calc['boron']
        if calc.get('sulphur'):        s.sulphur = calc['sulphur']
        if calc.get('zinc'):           s.zinc = calc['zinc']
        if calc.get('iron'):           s.iron = calc['iron']
        if calc.get('manganese'):      s.manganese = calc['manganese']
        if calc.get('copper'):         s.copper = calc['copper']
        sync_lab_tables(s)  # sets s.category AND populates TestResult / LabCalculation
        count += 1
    db.session.commit()
    return f'✅ Recalculated {count} samples! <a href="/">Go to Dashboard</a>'

# ── Multiplication Factors ──
@app.route('/dilution-factors')
@app.route('/multiplication-factors')
@staff_or_admin_required
def dilution_factors():
    factors = DilutionFactor.query.all()
    return render_template('dilution_factors.html', factors=factors)

@app.route('/dilution-factors/update', methods=['POST'])
@app.route('/multiplication-factors/update', methods=['POST'])
@admin_required
def update_dilution_factors():
    factors = DilutionFactor.query.all()
    for f in factors:
        new_value = request.form.get(f.parameter, type=float)
        if new_value is not None:
            f.factor = new_value
    db.session.commit()
    flash('✅ Multiplication factors updated successfully!', 'success')
    return redirect(url_for('dilution_factors'))

# ── Lab Calculation Wizard ──
@app.route('/lab-calculation')
@staff_or_admin_required
def lab_calculation():
    return render_template('lab_calculation.html', units=PARAMETER_UNITS)

def get_included_params(sample):
    """
    Returns a set of parameter keys included in the sample's test_package.
    If sample_source is govt or no test_package is set, returns all 12 parameters.
    """
    if not sample or sample.sample_source == 'govt' or not sample.test_package:
        return {'ph', 'ec', 'organic_carbon', 'nitrogen', 'phosphorus', 'potassium', 'sulphur', 'zinc', 'boron', 'iron', 'manganese', 'copper'}
    
    pkg = sample.test_package.lower()
    if '5.' in pkg or 'full' in pkg:
        return {'ph', 'ec', 'organic_carbon', 'nitrogen', 'phosphorus', 'potassium', 'sulphur', 'zinc', 'boron', 'iron', 'manganese', 'copper'}
        
    included = set()
    if '1.' in pkg or 'basic' in pkg:
        included.update({'ph', 'ec', 'organic_carbon'})
    if '2.' in pkg or 'major' in pkg:
        included.update({'nitrogen', 'phosphorus', 'potassium'})
    if '3.' in pkg or 'complete' in pkg:
        included.update({'ph', 'ec', 'organic_carbon', 'nitrogen', 'phosphorus', 'potassium'})
    if '4.' in pkg or 'micronutrient' in pkg:
        included.update({'zinc', 'iron', 'manganese', 'copper'})
        
    return included if included else {'ph', 'ec', 'organic_carbon', 'nitrogen', 'phosphorus', 'potassium', 'sulphur', 'zinc', 'boron', 'iron', 'manganese', 'copper'}

# ── API Routes ──
@app.route('/api/factors')
def api_factors():
    return jsonify(get_factors_dict())

@app.route('/api/units')
def api_units():
    return jsonify(PARAMETER_UNITS)

# ── UPDATED: returns ALL fields for auto-fill ──
@app.route('/api/sample-by-id/<sample_id>')
def api_sample_by_id(sample_id):
    sample = Sample.query.filter_by(sample_id=sample_id).first()
    if not sample:
        return jsonify({'found': False})
    
    inc_params = list(get_included_params(sample))
    return jsonify({
        'found': True,
        'test_package': sample.test_package,
        'included_params': inc_params,
        # Step 1
        'farmer_name':     sample.farmer_name,
        'village':         sample.village,
        'sample_type':     sample.sample_type,
        'collection_date': sample.collection_date,
        'phone_number':    sample.phone_number,
        'address':         sample.address,
        'survey_number':   sample.survey_number,
        'sample_source':   sample.sample_source,
        'scheme':          sample.scheme,
        # Step 2
        'ph':              sample.ph,
        'observed_ec':     sample.observed_ec,
        'ec_temperature':  sample.ec_temperature,
        'ec':              sample.ec,
        # Step 3
        'n_burette_a':     sample.n_burette_a,
        'n_burette_b':     sample.n_burette_b,
        'nitrogen':        sample.nitrogen,
        # Step 4
        'abs_phosphorus':  sample.abs_phosphorus,
        'phosphorus':      sample.phosphorus,
        'abs_potassium':   sample.abs_potassium,
        'potassium':       sample.potassium,
        # Step 5
        'abs_organic_carbon': sample.abs_organic_carbon,
        'organic_carbon':     sample.organic_carbon,
        'abs_boron':          sample.abs_boron,
        'boron':              sample.boron,
        'abs_sulphur':        sample.abs_sulphur,
        'sulphur':            sample.sulphur,
        # Step 6
        'abs_zinc':        sample.abs_zinc,
        'zinc':            sample.zinc,
        'abs_iron':        sample.abs_iron,
        'iron':            sample.iron,
        'abs_manganese':   sample.abs_manganese,
        'manganese':       sample.manganese,
        'abs_copper':      sample.abs_copper,
        'copper':          sample.copper,
        # Step 7
        'notes':           sample.notes,
        # Sign-off
        'analyzed_by':     sample.analyzed_by,
        'checked_by':      sample.checked_by,
        'approved_by':     sample.approved_by,
    })

@app.route('/api/save-sample', methods=['POST'])
def api_save_sample():
    """
    Kept for the Lab Calculation wizard's own save path
    (chemistry parameters only — basic info is set at /add time).
    """
    data = request.get_json()
    sample_id = data.get('sample_id')
    sample = Sample.query.filter_by(sample_id=sample_id).first() if sample_id else None

    ph = data.get('ph')
    nitrogen = data.get('nitrogen')
    phosphorus = data.get('phosphorus')
    potassium = data.get('potassium')

    if not sample:
        # Fallback: create a bare sample if none exists (shouldn't normally happen
        # since registration now happens via /add first)
        # NOTE: data.get('village', 'Unknown') would NOT catch this — the wizard always
        # sends a 'village' key, even when empty, so the key is present with value None/''.
        # .get()'s default only applies when the key is missing entirely, so we need `or` here.
        village = data.get('village') or 'Unknown'
        sample = Sample(sample_id=generate_sample_id(village), village=village)
        db.session.add(sample)
        db.session.flush()  # need sample.id before sync_lab_tables() can use it

    sample.ph                 = ph
    sample.ec                 = data.get('ec')
    sample.nitrogen           = nitrogen
    sample.phosphorus         = phosphorus
    sample.potassium          = potassium
    sample.iron               = data.get('iron')
    sample.manganese          = data.get('manganese')
    sample.copper             = data.get('copper')
    sample.zinc                = data.get('zinc')
    sample.boron               = data.get('boron')
    sample.organic_carbon      = data.get('organic_carbon')
    sample.sulphur              = data.get('sulphur')
    sample.observed_ec        = data.get('observed_ec')
    sample.ec_temperature     = data.get('ec_temperature')
    sample.n_burette_a        = data.get('n_burette_a')
    sample.n_burette_b        = data.get('n_burette_b')
    sample.abs_phosphorus     = data.get('abs_phosphorus')
    sample.abs_potassium      = data.get('abs_potassium')
    sample.abs_organic_carbon = data.get('abs_organic_carbon')
    sample.abs_boron          = data.get('abs_boron')
    sample.abs_sulphur        = data.get('abs_sulphur')
    sample.abs_zinc           = data.get('abs_zinc')
    sample.abs_iron           = data.get('abs_iron')
    sample.abs_manganese      = data.get('abs_manganese')
    sample.abs_copper         = data.get('abs_copper')
    sample.notes              = data.get('notes') or sample.notes
    sample.analyzed_by        = data.get('analyzed_by') or sample.analyzed_by
    sample.checked_by         = data.get('checked_by') or sample.checked_by
    sample.approved_by        = data.get('approved_by') or sample.approved_by

    if not sample.id:
        db.session.flush()
    sync_lab_tables(sample)  # sets sample.category AND populates TestResult / LabCalculation
    db.session.commit()
    return jsonify({'success': True, 'sample_id': sample.sample_id})

@app.route('/api/update-sample/<sample_id>', methods=['POST'])
def api_update_sample(sample_id):
    sample = Sample.query.filter_by(sample_id=sample_id).first()
    if not sample:
        return jsonify({'success': False, 'error': 'Sample not found'})
    data = request.get_json()
    sample.ph              = data.get('ph') or sample.ph
    sample.ec              = data.get('ec') or sample.ec
    sample.nitrogen        = data.get('nitrogen') or sample.nitrogen
    sample.phosphorus      = data.get('phosphorus') or sample.phosphorus
    sample.potassium       = data.get('potassium') or sample.potassium
    sample.iron            = data.get('iron') or sample.iron
    sample.manganese       = data.get('manganese') or sample.manganese
    sample.copper          = data.get('copper') or sample.copper
    sample.zinc             = data.get('zinc') or sample.zinc
    sample.boron            = data.get('boron') or sample.boron
    sample.organic_carbon   = data.get('organic_carbon') or sample.organic_carbon
    sample.sulphur           = data.get('sulphur') or sample.sulphur
    sample.observed_ec        = data.get('observed_ec') or sample.observed_ec
    sample.ec_temperature     = data.get('ec_temperature') or sample.ec_temperature
    sample.n_burette_a        = data.get('n_burette_a') or sample.n_burette_a
    sample.n_burette_b        = data.get('n_burette_b') or sample.n_burette_b
    sample.abs_phosphorus     = data.get('abs_phosphorus') or sample.abs_phosphorus
    sample.abs_potassium      = data.get('abs_potassium') or sample.abs_potassium
    sample.abs_organic_carbon = data.get('abs_organic_carbon') or sample.abs_organic_carbon
    sample.abs_boron          = data.get('abs_boron') or sample.abs_boron
    sample.abs_sulphur        = data.get('abs_sulphur') or sample.abs_sulphur
    sample.abs_zinc           = data.get('abs_zinc') or sample.abs_zinc
    sample.abs_iron           = data.get('abs_iron') or sample.abs_iron
    sample.abs_manganese      = data.get('abs_manganese') or sample.abs_manganese
    sample.abs_copper         = data.get('abs_copper') or sample.abs_copper
    sample.farmer_name        = data.get('farmer_name') or sample.farmer_name
    sample.village            = data.get('village') or sample.village
    sample.sample_type        = data.get('sample_type') or sample.sample_type
    sample.collection_date    = data.get('collection_date') or sample.collection_date
    sample.phone_number       = data.get('phone_number') or sample.phone_number
    sample.address            = data.get('address') or sample.address
    sample.survey_number      = data.get('survey_number') or sample.survey_number
    sample.sample_source      = data.get('sample_source') or sample.sample_source
    sample.scheme              = data.get('scheme') or sample.scheme
    sample.notes               = data.get('notes') or sample.notes
    sample.analyzed_by         = data.get('analyzed_by') or sample.analyzed_by
    sample.checked_by          = data.get('checked_by') or sample.checked_by
    sample.approved_by         = data.get('approved_by') or sample.approved_by
    sync_lab_tables(sample)  # sets sample.category AND populates TestResult / LabCalculation
    db.session.commit()
    return jsonify({'success': True, 'sample_id': sample.sample_id})

def get_param_5level_rating(param, val):
    if val is None:
        return None
    try:
        val = float(val)
    except (ValueError, TypeError):
        return None

    # 5 Glossy Vector SVGs from User Reference HTML:
    SVG_GREEN = '<svg viewBox="0 0 100 100" width="100%" height="100%"><defs><radialGradient id="gGreen" cx="35%" cy="30%" r="75%"><stop offset="0%" stop-color="#86EFAC"/><stop offset="55%" stop-color="#22C55E"/><stop offset="100%" stop-color="#16803C"/></radialGradient></defs><circle cx="50" cy="50" r="47" fill="url(#gGreen)"/><ellipse cx="38" cy="28" rx="20" ry="12" fill="#ffffff" opacity="0.28"/><circle cx="34" cy="46" r="5.5" fill="#1a1a1a"/><circle cx="66" cy="46" r="5.5" fill="#1a1a1a"/><path d="M30 60 Q50 82 70 60 Q68 76 50 76 Q32 76 30 60 Z" fill="#7A1F1F"/><path d="M33 61 Q50 72 67 61" stroke="#ffffff" stroke-width="4" fill="none" stroke-linecap="round"/></svg>'
    SVG_YELLOW = '<svg viewBox="0 0 100 100" width="100%" height="100%"><defs><radialGradient id="gYellow" cx="35%" cy="30%" r="75%"><stop offset="0%" stop-color="#FDE68A"/><stop offset="55%" stop-color="#EAB308"/><stop offset="100%" stop-color="#A5730A"/></radialGradient></defs><circle cx="50" cy="50" r="47" fill="url(#gYellow)"/><ellipse cx="38" cy="28" rx="20" ry="12" fill="#ffffff" opacity="0.3"/><circle cx="34" cy="46" r="5" fill="#1a1a1a"/><circle cx="66" cy="46" r="5" fill="#1a1a1a"/><path d="M32 62 Q50 74 68 62" stroke="#1a1a1a" stroke-width="4.5" fill="none" stroke-linecap="round"/></svg>'
    SVG_ORANGE = '<svg viewBox="0 0 100 100" width="100%" height="100%"><defs><radialGradient id="gOrange" cx="35%" cy="30%" r="75%"><stop offset="0%" stop-color="#FDBA74"/><stop offset="55%" stop-color="#F97316"/><stop offset="100%" stop-color="#B4530A"/></radialGradient></defs><circle cx="50" cy="50" r="47" fill="url(#gOrange)"/><ellipse cx="38" cy="28" rx="20" ry="12" fill="#ffffff" opacity="0.28"/><circle cx="34" cy="46" r="5" fill="#1a1a1a"/><circle cx="66" cy="46" r="5" fill="#1a1a1a"/><line x1="33" y1="65" x2="67" y2="65" stroke="#1a1a1a" stroke-width="4.5" stroke-linecap="round"/></svg>'
    SVG_RED = '<svg viewBox="0 0 100 100" width="100%" height="100%"><defs><radialGradient id="gRed" cx="35%" cy="30%" r="75%"><stop offset="0%" stop-color="#FCA5A5"/><stop offset="55%" stop-color="#EF4444"/><stop offset="100%" stop-color="#9B1C1C"/></radialGradient></defs><circle cx="50" cy="50" r="47" fill="url(#gRed)"/><ellipse cx="38" cy="28" rx="20" ry="12" fill="#ffffff" opacity="0.25"/><line x1="26" y1="38" x2="40" y2="44" stroke="#1a1a1a" stroke-width="4" stroke-linecap="round"/><line x1="74" y1="38" x2="60" y2="44" stroke="#1a1a1a" stroke-width="4" stroke-linecap="round"/><circle cx="34" cy="50" r="5" fill="#1a1a1a"/><circle cx="66" cy="50" r="5" fill="#1a1a1a"/><path d="M32 70 Q50 60 68 70 Q50 66 32 70 Z" fill="#1a1a1a"/></svg>'
    SVG_PURPLE = '<svg viewBox="0 0 100 100" width="100%" height="100%"><defs><radialGradient id="gPurple" cx="35%" cy="30%" r="75%"><stop offset="0%" stop-color="#D8B4FE"/><stop offset="55%" stop-color="#9333EA"/><stop offset="100%" stop-color="#5B1A8C"/></radialGradient></defs><circle cx="50" cy="50" r="47" fill="url(#gPurple)"/><ellipse cx="38" cy="28" rx="20" ry="12" fill="#ffffff" opacity="0.22"/><line x1="24" y1="40" x2="42" y2="47" stroke="#1a1a1a" stroke-width="4.5" stroke-linecap="round"/><line x1="76" y1="40" x2="58" y2="47" stroke="#1a1a1a" stroke-width="4.5" stroke-linecap="round"/><circle cx="34" cy="52" r="5.5" fill="#1a1a1a"/><circle cx="66" cy="52" r="5.5" fill="#1a1a1a"/><path d="M30 64 Q50 60 70 64 Q68 78 50 78 Q32 78 30 64 Z" fill="#1a1a1a"/><line x1="38" y1="66" x2="38" y2="73" stroke="#D8B4FE" stroke-width="2.5"/><line x1="46" y1="65" x2="46" y2="75" stroke="#D8B4FE" stroke-width="2.5"/><line x1="54" y1="65" x2="54" y2="75" stroke="#D8B4FE" stroke-width="2.5"/><line x1="62" y1="66" x2="62" y2="73" stroke="#D8B4FE" stroke-width="2.5"/></svg>'

    C_LOWER    = {'level': 'lower',    'label': 'LOWER',    'color': '#22C55E', 'bg': '#DCFCE7', 'border': '#16803C', 'svg': SVG_GREEN}
    C_LOW      = {'level': 'low',      'label': 'LOW',      'color': '#EAB308', 'bg': '#FEF3C7', 'border': '#A5730A', 'svg': SVG_YELLOW}
    C_MODERATE = {'level': 'moderate', 'label': 'MODERATE', 'color': '#F97316', 'bg': '#FFEDD5', 'border': '#B4530A', 'svg': SVG_ORANGE}
    C_HIGH     = {'level': 'high',     'label': 'HIGH',     'color': '#EF4444', 'bg': '#FEE2E2', 'border': '#9B1C1C', 'svg': SVG_RED}
    C_HIGHER   = {'level': 'higher',   'label': 'HIGHER',   'color': '#9333EA', 'bg': '#F3E8FF', 'border': '#5B1A8C', 'svg': SVG_PURPLE}

    if param == 'ph':
        range_str = "6.5 - 8.5"
        if val < 5.5:
            return {**C_LOWER, 'range': range_str}
        elif val < 6.5:
            return {**C_LOW, 'range': range_str}
        elif val <= 7.5:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 8.5:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'ec':
        range_str = "< 1.0"
        if val < 0.2:
            return {**C_LOWER, 'range': range_str}
        elif val <= 0.8:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 1.6:
            return {**C_HIGH, 'range': range_str}
        elif val <= 2.5:
            return {**C_LOW, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'organic_carbon':
        range_str = "0.51 - 0.75 %"
        if val < 0.25:
            return {**C_LOWER, 'range': range_str}
        elif val <= 0.50:
            return {**C_LOW, 'range': range_str}
        elif val <= 0.75:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 1.0:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'nitrogen':
        range_str = "280 - 560"
        if val < 140:
            return {**C_LOWER, 'range': range_str}
        elif val < 280:
            return {**C_LOW, 'range': range_str}
        elif val <= 420:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 560:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'phosphorus':
        range_str = "10 - 25"
        if val < 5:
            return {**C_LOWER, 'range': range_str}
        elif val < 10:
            return {**C_LOW, 'range': range_str}
        elif val <= 17.5:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 25:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'potassium':
        range_str = "120 - 280"
        if val < 60:
            return {**C_LOWER, 'range': range_str}
        elif val < 120:
            return {**C_LOW, 'range': range_str}
        elif val <= 200:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 280:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'sulphur':
        range_str = "10 - 20"
        if val < 5:
            return {**C_LOWER, 'range': range_str}
        elif val < 10:
            return {**C_LOW, 'range': range_str}
        elif val <= 15:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 20:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'zinc':
        range_str = "> 0.6"
        if val < 0.3:
            return {**C_LOWER, 'range': range_str}
        elif val < 0.6:
            return {**C_LOW, 'range': range_str}
        elif val <= 1.0:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 1.5:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'boron':
        range_str = "> 0.5"
        if val < 0.25:
            return {**C_LOWER, 'range': range_str}
        elif val < 0.5:
            return {**C_LOW, 'range': range_str}
        elif val <= 0.75:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 1.0:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'iron':
        range_str = "> 4.5"
        if val < 2.5:
            return {**C_LOWER, 'range': range_str}
        elif val < 4.5:
            return {**C_LOW, 'range': range_str}
        elif val <= 7.5:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 10.0:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'manganese':
        range_str = "> 2.0"
        if val < 1.0:
            return {**C_LOWER, 'range': range_str}
        elif val < 2.0:
            return {**C_LOW, 'range': range_str}
        elif val <= 3.5:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 5.0:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    elif param == 'copper':
        range_str = "> 0.2"
        if val < 0.1:
            return {**C_LOWER, 'range': range_str}
        elif val < 0.2:
            return {**C_LOW, 'range': range_str}
        elif val <= 0.5:
            return {**C_MODERATE, 'range': range_str}
        elif val <= 0.8:
            return {**C_HIGH, 'range': range_str}
        else:
            return {**C_HIGHER, 'range': range_str}

    return None

# ── Official Recommended Fertilizer Dose Dataset (From Department Excel) ──
CROP_RECOMMENDED_DOSES = {
    # Fruit Crops (N, P, K per plant, FYM in kg/plant)
    'Mango 1st year': {'n': 0.300, 'p': 0.300, 'k': 0.100, 'fym': 10, 'unit': 'Grams Per Plant'},
    'Mango 2nd year': {'n': 0.600, 'p': 0.600, 'k': 0.200, 'fym': 20, 'unit': 'Grams Per Plant'},
    'Mango 3rd year': {'n': 0.900, 'p': 0.900, 'k': 0.300, 'fym': 30, 'unit': 'Grams Per Plant'},
    'Mango 4th year': {'n': 1.200, 'p': 1.200, 'k': 0.400, 'fym': 40, 'unit': 'Grams Per Plant'},
    'Mango 5th year': {'n': 1.500, 'p': 1.500, 'k': 0.500, 'fym': 50, 'unit': 'Grams Per Plant'},
    'Mango 6th year': {'n': 1.800, 'p': 1.800, 'k': 0.600, 'fym': 60, 'unit': 'Grams Per Plant'},
    'Mango 7th year': {'n': 2.100, 'p': 2.100, 'k': 0.700, 'fym': 70, 'unit': 'Grams Per Plant'},
    'Mango 8th year': {'n': 2.400, 'p': 2.400, 'k': 0.800, 'fym': 80, 'unit': 'Grams Per Plant'},
    'Mango 9th year': {'n': 2.700, 'p': 2.700, 'k': 0.900, 'fym': 90, 'unit': 'Grams Per Plant'},
    'Mango 10th year onward': {'n': 3.000, 'p': 3.000, 'k': 1.000, 'fym': 100, 'unit': 'Grams Per Plant'},

    'Cashew 1st year': {'n': 0.250, 'p': 0.630, 'k': 0.630, 'fym': 10, 'unit': 'Grams Per Plant'},
    'Cashew 2nd year': {'n': 0.500, 'p': 0.125, 'k': 0.125, 'fym': 20, 'unit': 'Grams Per Plant'},
    'Cashew 3rd year': {'n': 0.750, 'p': 0.188, 'k': 0.188, 'fym': 30, 'unit': 'Grams Per Plant'},
    'Cashew 4th year Onwards': {'n': 1.000, 'p': 0.250, 'k': 0.250, 'fym': 40, 'unit': 'Grams Per Plant'},

    'Coconut 1st year': {'n': 0.200, 'p': 0.100, 'k': 0.200, 'fym': 10, 'unit': 'Grams Per Plant'},
    'Coconut 2nd year': {'n': 0.400, 'p': 0.200, 'k': 0.400, 'fym': 20, 'unit': 'Grams Per Plant'},
    'Coconut 3rd year': {'n': 0.600, 'p': 0.300, 'k': 0.600, 'fym': 30, 'unit': 'Grams Per Plant'},
    'Coconut 4th year': {'n': 0.800, 'p': 0.400, 'k': 0.800, 'fym': 40, 'unit': 'Grams Per Plant'},
    'Coconut 5th year Onwards': {'n': 1.000, 'p': 0.500, 'k': 1.000, 'fym': 50, 'unit': 'Grams Per Plant'},

    'Chikku 1st year': {'n': 0.150, 'p': 0.150, 'k': 0.150, 'fym': 10, 'unit': 'Grams Per Plant'},
    'Chikku 2nd year': {'n': 0.300, 'p': 0.300, 'k': 0.300, 'fym': 20, 'unit': 'Grams Per Plant'},
    'Chikku 5th year': {'n': 0.750, 'p': 0.750, 'k': 0.750, 'fym': 50, 'unit': 'Grams Per Plant'},
    'Chikku 10th year': {'n': 1.500, 'p': 1.500, 'k': 1.500, 'fym': 100, 'unit': 'Grams Per Plant'},
    'Chikku 20th year Onwards': {'n': 3.000, 'p': 3.000, 'k': 3.000, 'fym': 200, 'unit': 'Grams Per Plant'},

    'Arecanut 1st year': {'n': 0.050, 'p': 0.050, 'k': 0.050, 'fym': 10, 'unit': 'Grams Per Plant'},
    'Arecanut 2nd year': {'n': 0.100, 'p': 0.100, 'k': 0.100, 'fym': 20, 'unit': 'Grams Per Plant'},
    'Arecanut 3rd year Onwards': {'n': 0.150, 'p': 0.150, 'k': 0.150, 'fym': 30, 'unit': 'Grams Per Plant'},

    'Black Pepper 1st year': {'n': 0.050, 'p': 0.025, 'k': 0.050, 'fym': 3, 'unit': 'Grams Per Plant'},
    'Black Pepper 3rd year Onwards': {'n': 0.150, 'p': 0.075, 'k': 0.150, 'fym': 10, 'unit': 'Grams Per Plant'},

    'Nutmeg 1st year': {'n': 0.050, 'p': 0.025, 'k': 0.100, 'fym': 1, 'unit': 'Grams Per Plant'},
    'Nutmeg 10th year onward': {'n': 0.500, 'p': 0.250, 'k': 1.000, 'fym': 10, 'unit': 'Grams Per Plant'},

    'Cinnamon 1st year': {'n': 0.020, 'p': 0.020, 'k': 0.020, 'fym': 1, 'unit': 'Grams Per Plant'},
    'Cinnamon 10th year onward': {'n': 0.200, 'p': 0.200, 'k': 0.200, 'fym': 20, 'unit': 'Grams Per Plant'},

    'Kokum 1st year': {'n': 0.045, 'p': 0.025, 'k': 0.030, 'fym': 2, 'unit': 'Grams Per Plant'},
    'Kokum 10th year onward': {'n': 0.450, 'p': 0.250, 'k': 0.300, 'fym': 20, 'unit': 'Grams Per Plant'},

    # Vegetables
    'Brinjal': {'n': 150, 'p': 50, 'k': 50, 'fym': 20000, 'unit': 'Kg Per Hectare'},
    'Chilli': {'n': 150, 'p': 50, 'k': 50, 'fym': 15000, 'unit': 'Kg Per Hectare'},
    'Okra': {'n': 100, 'p': 50, 'k': 25, 'fym': 15000, 'unit': 'Kg Per Hectare'},
    'Ridge Gourds': {'n': 100, 'p': 50, 'k': 50, 'fym': 15000, 'unit': 'Kg Per Hectare'},
    'Cabbage': {'n': 120, 'p': 60, 'k': 60, 'fym': 20000, 'unit': 'Kg Per Hectare'},
    'Bitter Gourds': {'n': 120, 'p': 60, 'k': 30, 'fym': 15000, 'unit': 'Kg Per Hectare'},

    # Pulses
    'Cow Pea': {'n': 60, 'p': 50, 'k': 50, 'fym': 20, 'unit': 'Kg Per Hectare'},

    # Cereals
    'Paddy': {'n': 100, 'p': 50, 'k': 50, 'fym': 7.5, 'unit': 'Kg Per Hectare'},
    'Paddy (Hybrid)': {'n': 150, 'p': 50, 'k': 50, 'fym': 7.5, 'unit': 'Kg Per Hectare'},
    'Nachani': {'n': 80, 'p': 40, 'k': 40, 'fym': 5, 'unit': 'Kg Per Hectare'},
}

def get_detailed_recommendations(sample, ratings):
    raw_crop = (sample.crop or '').strip()
    if not raw_crop or raw_crop in ['—', '-', 'None', 'Null', '']:
        return {
            'has_crop': False,
            'crop': 'Not Specified',
            'grd_summary': {},
            'comb1': [],
            'comb2': [],
            'organic': []
        }

    crop_name = raw_crop
    # Check if crop exists in dataset or fuzzy match
    dose = CROP_RECOMMENDED_DOSES.get(crop_name)
    if not dose:
        # Default fallback match for Mango / Cashew / Paddy
        for k in CROP_RECOMMENDED_DOSES:
            if k.lower() in crop_name.lower() or crop_name.lower() in k.lower():
                dose = CROP_RECOMMENDED_DOSES[k]
                break
    if not dose:
        dose = CROP_RECOMMENDED_DOSES['Paddy']

    n_val = sample.nitrogen
    p_val = sample.phosphorus
    k_val = sample.potassium

    n_rating = ratings.get('nitrogen')
    p_rating = ratings.get('phosphorus')
    k_rating = ratings.get('potassium')
    
    # ── GRD Direct Comparison Logic ──
    # If tested lab value > target GRD (or rating is HIGH/HIGHER): Dose REDUCED (0.75x or 0.5x)
    # If tested lab value == target GRD (or rating is MODERATE): Dose STANDARD (1.0x OK)
    # If tested lab value < target GRD (or rating is LOW/LOWER): Dose INCREASED (1.25x to replenish soil)
    
    # Nitrogen multiplier
    if n_val is not None and n_val > dose['n']:
        mult_n = 0.75  # Reduced dose for excess Nitrogen
        n_status = f"{n_val} kg/ha (Higher than GRD {dose['n']}) → Reduced Dose (-25%)"
    elif n_rating and n_rating['level'] in ['high', 'higher']:
        mult_n = 0.75
        n_status = f"HIGH → Reduced Dose (-25%)"
    elif n_val is not None and n_val < (dose['n'] * 0.85):
        mult_n = 1.25  # Increased dose for deficient Nitrogen
        n_status = f"{n_val} kg/ha (Lower than GRD {dose['n']}) → Increased Dose (+25%)"
    elif n_rating and n_rating['level'] in ['low', 'lower']:
        mult_n = 1.25
        n_status = f"LOW → Increased Dose (+25%)"
    else:
        mult_n = 1.0   # Standard GRD dose OK
        n_status = f"{n_val if n_val is not None else 'OK'} (Sufficient / Matches GRD {dose['n']}) → Standard Dose (100%)"

    # Phosphorus multiplier
    if p_val is not None and p_val > dose['p']:
        mult_p = 0.75
        p_status = f"{p_val} kg/ha (Higher than GRD {dose['p']}) → Reduced Dose (-25%)"
    elif p_rating and p_rating['level'] in ['high', 'higher']:
        mult_p = 0.75
        p_status = f"HIGH → Reduced Dose (-25%)"
    elif p_val is not None and p_val < (dose['p'] * 0.85):
        mult_p = 1.25
        p_status = f"{p_val} kg/ha (Lower than GRD {dose['p']}) → Increased Dose (+25%)"
    elif p_rating and p_rating['level'] in ['low', 'lower']:
        mult_p = 1.25
        p_status = f"LOW → Increased Dose (+25%)"
    else:
        mult_p = 1.0
        p_status = f"{p_val if p_val is not None else 'OK'} (Sufficient / Matches GRD {dose['p']}) → Standard Dose (100%)"

    # Potassium multiplier
    if k_val is not None and k_val > dose['k']:
        mult_k = 0.75
        k_status = f"{k_val} kg/ha (Higher than GRD {dose['k']}) → Reduced Dose (-25%)"
    elif k_rating and k_rating['level'] in ['high', 'higher']:
        mult_k = 0.75
        k_status = f"HIGH → Reduced Dose (-25%)"
    elif k_val is not None and k_val < (dose['k'] * 0.85):
        mult_k = 1.25
        k_status = f"{k_val} kg/ha (Lower than GRD {dose['k']}) → Increased Dose (+25%)"
    elif k_rating and k_rating['level'] in ['low', 'lower']:
        mult_k = 1.25
        k_status = f"LOW → Increased Dose (+25%)"
    else:
        mult_k = 1.0
        k_status = f"{k_val if k_val is not None else 'OK'} (Sufficient / Matches GRD {dose['k']}) → Standard Dose (100%)"

    dap_amt = int(round(dose['p'] * mult_p * 1000)) if 'Grams' in dose['unit'] else int(round(dose['p'] * mult_p))
    mop_amt = int(round(dose['k'] * mult_k * 1000)) if 'Grams' in dose['unit'] else int(round(dose['k'] * mult_k))
    urea_amt = int(round(dose['n'] * mult_n * 1000 * 2.17)) if 'Grams' in dose['unit'] else int(round(dose['n'] * mult_n * 2.17))
    ssp_amt = int(round(dose['p'] * mult_p * 1000 * 3.0)) if 'Grams' in dose['unit'] else int(round(dose['p'] * mult_p * 3.0))

    comb1 = [
        f"DAP - {dap_amt} {dose['unit']}",
        f"MOP - {mop_amt} {dose['unit']}",
        f"Urea - {urea_amt} {dose['unit']}"
    ]
    comb2 = [
        f"SSP - {ssp_amt} {dose['unit']}",
        f"MOP - {mop_amt} {dose['unit']}",
        f"Urea - {urea_amt} {dose['unit']}"
    ]
    fym_str = f"FYM: {dose['fym']} {'Tons Per Hectare' if isinstance(dose['fym'], float) else ('Kg Per Hectare' if 'Hectare' in dose['unit'] else 'Kg Per Plant')}"
    organic = [
        fym_str,
        "Oil Cake (Neem/Castor/Karanja): 500-750 Kg Per Hectare",
        "Bio-fertilizers: Azotobacter / Azospirillum 10 Kg/ha + PSB @ 10 kg/ha + VAM (AMF) @ 10-12.5 kg/ha",
        "Method: Basal application in split doses",
        "Rock phosphate: 200-300 Kg Per Hectare"
    ]
    return {
        'has_crop': True,
        'crop': crop_name,
        'grd_target': dose,
        'status_n': n_status,
        'status_p': p_status,
        'status_k': k_status,
        'comb1': comb1,
        'comb2': comb2,
        'organic': organic
    }

# ── Soil Health Card ──
@app.route('/soil-health-card/<int:id>')
@login_required
def soil_health_card(id):
    sample = db.session.get(Sample, id)
    if not sample:
        return redirect(url_for('all_samples'))
    included_params = get_included_params(sample)
    
    # Calculate 5-level parameter ratings with emojis
    param_ratings = {}
    for p in ['ph', 'ec', 'organic_carbon', 'nitrogen', 'phosphorus', 'potassium', 'sulphur', 'zinc', 'boron', 'iron', 'manganese', 'copper']:
        val = getattr(sample, p, None)
        if val is not None:
            param_ratings[p] = get_param_5level_rating(p, val)

    recommendations = get_detailed_recommendations(sample, param_ratings)

    return render_template(
        'soil_health_card.html',
        sample=sample,
        units=PARAMETER_UNITS,
        included_params=included_params,
        param_ratings=param_ratings,
        recommendations=recommendations
    )

# ── Users ──
@app.route('/users')
@app.route('/User%20Management')
@app.route('/User Management')
@app.route('/user-management')
@admin_required
def users():
    all_users = User.query.all()
    return render_template('users.html', users=all_users)

@app.route('/delete_user/<int:id>')
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully!', 'success')
    return redirect(url_for('users'))

@app.route('/reset_user_password/<int:id>', methods=['POST'])
@admin_required
def reset_user_password(id):
    user = User.query.get_or_404(id)
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    if not new_password:
        flash('Password cannot be empty.', 'error')
        return redirect(url_for('users'))

    if new_password != confirm_password:
        flash('Passwords do not match. Please try again.', 'error')
        return redirect(url_for('users'))

    user.password = generate_password_hash(new_password)
    user.failed_attempts = 0
    user.last_failed_at = None
    db.session.commit()

    flash(f'✅ Password reset successfully for user "{user.username}"!', 'success')
    return redirect(url_for('users'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/bill/<int:id>')
@login_required
def generate_bill(id):
    sample = db.session.get(Sample, id)
    if not sample:
        sample = Sample.query.first()
    if not sample:
        flash("No sample found.", "error")
        return redirect(url_for('all_samples'))
    return render_template('bill_receipt.html', sample=sample)

try:
    with app.app_context():
        db.create_all()
        seed_dilution_factors()
        seed_admin_account()
except Exception as e:
    print(f"Startup DB init log: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
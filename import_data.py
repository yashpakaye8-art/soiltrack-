import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app import app, db, Sample
import openpyxl
import shutil

def get_category(ph, nitrogen, phosphorus, potassium):
    score = 0
    if ph and 6.0 <= ph <= 7.5:
        score += 3
    elif ph and 5.5 <= ph < 6.0:
        score += 2
    if nitrogen:
        if nitrogen > 280: score += 3
        elif nitrogen >= 140: score += 2
    if phosphorus:
        if phosphorus > 40: score += 3
        elif phosphorus >= 20: score += 2
    if potassium:
        if potassium > 160: score += 3
        elif potassium >= 80: score += 2
    if score >= 10:
        return "Fertile"
    elif score >= 6:
        return "Moderate"
    else:
        return "Poor"

def import_data():
    wb_analysis = openpyxl.load_workbook(
        'Soil Data Analysis_Final_Uploading_Jambhulwadi.xlsx',
        data_only=True
    )
    wb_govt = openpyxl.load_workbook(
        'RATNAGIRI__2025-26_660f941a5c8405ca8375c7c6_batch_1.xlsx',
        data_only=True
    )

    ws_cumulative = wb_analysis['cumulative']
    ws_govt = wb_govt['Mysheet1']

    farmers = []
    for row in ws_govt.iter_rows(min_row=2, values_only=True):
        if row[0]:
            village_raw = str(row[4]) if row[4] else 'Unknown'
            village = village_raw.split(' - ')[0].strip()
            farmers.append({
                'test_id': row[0],
                'farmer_name': row[1],
                'phone': row[2],
                'survey_no': row[3],
                'village': village,
            })

    print(f"Found {len(farmers)} farmers in govt file")

    samples_added = 0
    for i, row in enumerate(ws_cumulative.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue

        sample_no  = int(row[0])
        ph         = row[1]
        ec         = row[2]
        oc         = row[3]
        nitrogen   = row[4]
        phosphorus = row[5]
        potassium  = row[6]
        sulphur    = row[7]
        zinc       = row[8]
        boron      = row[9]
        iron       = row[10]
        manganese  = row[11]
        copper     = row[12]

        farmer_name  = farmers[i]['farmer_name'] if i < len(farmers) else f'Farmer {sample_no}'
        village      = farmers[i]['village'] if i < len(farmers) else 'Jambhulwadi'
        village_code = village[:3].upper()

        count = Sample.query.filter(
            Sample.sample_id.like(f"{village_code}%")
        ).count()
        sample_id = f"{village_code}{str(count + 1).zfill(2)}"

        while Sample.query.filter_by(sample_id=sample_id).first():
            count += 1
            sample_id = f"{village_code}{str(count + 1).zfill(2)}"

        category = get_category(ph, nitrogen, phosphorus, potassium)

        new_sample = Sample(
            sample_id       = sample_id,
            village         = village,
            sample_type     = 'Government',
            farmer_name     = farmer_name,
            collection_date = '2025-2026',
            ph              = ph,
            ec              = ec,
            nitrogen        = nitrogen,
            phosphorus      = phosphorus,
            potassium       = potassium,
            iron            = iron,
            manganese       = manganese,
            copper          = copper,
            zinc            = zinc,
            boron           = boron,
            organic_carbon  = oc,
            sulphur         = sulphur,
            temperature     = None,
            moisture        = None,
            notes           = 'Imported from govt portal',
            category        = category
        )

        db.session.add(new_sample)
        samples_added += 1
        print(f"✅ {sample_id} — {farmer_name} — {village} — {category}")

    db.session.commit()
    print(f"\n🎉 Successfully imported {samples_added} samples!")

if __name__ == '__main__':
    src1 = r'C:\soil sample testing FP\Soil Data Analysis_Final_Uploading_Jambhulwadi.xlsx'
    src2 = r'C:\soil sample testing FP\RATNAGIRI__2025-26_660f941a5c8405ca8375c7c6_batch_1.xlsx'

    if os.path.exists(src1):
        shutil.copy(src1, 'Soil Data Analysis_Final_Uploading_Jambhulwadi.xlsx')
        print("✅ Copied analysis file")
    else:
        print("❌ Analysis file not found! Check the path.")

    if os.path.exists(src2):
        shutil.copy(src2, 'RATNAGIRI__2025-26_660f941a5c8405ca8375c7c6_batch_1.xlsx')
        print("✅ Copied govt file")
    else:
        print("❌ Govt file not found! Check the path.")

    with app.app_context():
        import_data()